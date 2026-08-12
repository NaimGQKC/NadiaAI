"""Export heir leads to a client-ready Excel workbook.

Honest by construction — the columns never imply data we don't have:

  * HEIR email vs OFFICE email are SEPARATE columns. A court/notary address
    (contact_source 'juzgado'/'notaria') is the office handling the estate, NOT the
    heir's own address; showing it in the heir Email column invites someone to write
    to a courthouse as if it were the family.
  * Email confidence is shown only for an actual heir email.
  * Property address is blanked unless it is street-level (contains a number) — a bare
    city name is not an address.
  * Town is stripped of obituary-scrape boilerplate ("Zaragoza Ayer Día" -> "Zaragoza").

Read-only: never mutates the DB, spends no credits. Produces a formatted .xlsx when
openpyxl is available, else a UTF-8 CSV Excel opens cleanly.

Usage (on the runner, where the persistent DB lives):
    python tools/export_leads.py
    HEIR_EXPORT_CCAA=zaragoza HEIR_EXPORT_EMAILED_ONLY=1 python tools/export_leads.py

Env:
    HEIR_EXPORT_CCAA          scope: CCAA ("aragon") or province ("zaragoza"); blank = all.
    HEIR_EXPORT_TIERS         comma list of tiers; default "A,B".
    HEIR_EXPORT_EMAILED_ONLY  "1" = only rows with a HEIR email; default "0".
    HEIR_EXPORT_CONF          "high" = only high-confidence heir emails; default all.
    HEIR_EXPORT_OUT           output path; default ./nadia_leads_export.xlsx
    NADIA_DB_PATH             persistent DB path.
"""

from __future__ import annotations

import csv
import os
import re
import sqlite3
import sys

sys.path.insert(0, os.path.dirname(__file__))  # tools/  (for test_heir_phones helpers)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
from nadia_ai.config import DB_PATH  # noqa: E402

# Reuse the exact scope / name-quality tests the enrichment uses, so the export scope
# matches the run scope and no junk names reach the client.
import test_heir_phones as tp  # noqa: E402

CCAA = os.getenv("HEIR_EXPORT_CCAA", "aragon").strip().lower()
TIERS = [t.strip().upper() for t in os.getenv("HEIR_EXPORT_TIERS", "A,B").split(",") if t.strip()]
EMAILED_ONLY = os.getenv("HEIR_EXPORT_EMAILED_ONLY", "0").strip() == "1"
CONF = (os.getenv("HEIR_EXPORT_CONF") or "").strip().lower()
OUT = os.getenv("HEIR_EXPORT_OUT") or os.path.join(os.getcwd(), "nadia_leads_export.xlsx")

# Columns read from `leads` (only those that exist are fetched). Some are logic-only.
_FETCH = [
    "causante", "heir_name", "localidad", "region", "tier",
    "contact_email", "contact_confidence", "contact_source",
    "date_of_death", "referencia_catastral", "direccion",
]

_OFFICE_SOURCES = {"juzgado", "notaria"}
# Obituary scrapes sometimes append the notice's dateline to the town
# ("Zaragoza Ayer Día 3..."). Cut the town at that boilerplate.
_TOWN_NOISE = re.compile(r"(?i)\s+(ayer|hoy|d[ií]a|falleci|el\s+pasado)\b.*$")


def _g(r, col: str):
    return r[col] if col in r.keys() else None


def _is_office(r) -> bool:
    return (_g(r, "contact_source") or "").strip().lower() in _OFFICE_SOURCES


def _heir_email(r) -> str:
    """The heir's own email — never an office address."""
    em = (_g(r, "contact_email") or "").strip()
    return "" if (not em or _is_office(r)) else em


def _office_email(r) -> str:
    em = (_g(r, "contact_email") or "").strip()
    return em if (em and _is_office(r)) else ""


def _town(r) -> str:
    return _TOWN_NOISE.sub("", (_g(r, "localidad") or "").strip()).strip()


def _property_address(r) -> str:
    """Only a real street-level address. A bare city name is not an address."""
    a = (_g(r, "direccion") or "").strip()
    return a if any(ch.isdigit() for ch in a) else ""


def _confidence(r) -> str:
    """Confidence belongs to a heir email; blank when there isn't one (stale labels
    linger in the DB after a re-clear and would otherwise read as a real result)."""
    return (_g(r, "contact_confidence") or "").strip() if _heir_email(r) else ""


# Output spec: (header, value function). Order is the client's reading order.
_OUTPUT = [
    ("Deceased", lambda r: _g(r, "causante")),
    ("Heir", lambda r: _g(r, "heir_name")),
    ("Town", _town),
    ("Region", lambda r: _g(r, "region")),
    ("Tier", lambda r: _g(r, "tier")),
    ("Heir email", _heir_email),
    ("Email confidence", _confidence),
    ("Office email (court/notary)", _office_email),
    ("Date of death", lambda r: _g(r, "date_of_death")),
    ("Catastro ref", lambda r: _g(r, "referencia_catastral")),
    ("Property address", _property_address),
]


def _fetch(conn: sqlite3.Connection) -> list:
    have = {r[1] for r in conn.execute("PRAGMA table_info(leads)").fetchall()}
    cols = [c for c in _FETCH if c in have]
    if not cols:
        return []
    tier_ph = ",".join("?" for _ in TIERS)
    where = ["heir_name IS NOT NULL AND heir_name != ''"]
    if TIERS:
        where.append(f"tier IN ({tier_ph})")
    rows = conn.execute(
        f"SELECT {', '.join(cols)} FROM leads WHERE {' AND '.join(where)}",
        tuple(TIERS),
    ).fetchall()

    def keep(r) -> bool:
        if CCAA and not tp._in_scope(_g(r, "region") or "", _g(r, "localidad") or "", CCAA):
            return False
        if not tp._good_heir_name(_g(r, "heir_name") or ""):
            return False
        if EMAILED_ONLY and not _heir_email(r):
            return False
        if CONF == "high" and "high" not in _confidence(r).lower():
            return False
        return True

    def rank(t):
        i, r = t
        conf = _confidence(r).lower()
        # Heir emails first (high before medium), then office-email rows, then the rest.
        if _heir_email(r):
            bucket, sub = 0, (0 if "high" in conf else 1)
        elif _office_email(r):
            bucket, sub = 1, 0
        else:
            bucket, sub = 2, 0
        return (bucket, sub, _g(r, "region") or "", _town(r), i)

    return [r for _i, r in sorted(((i, r) for i, r in enumerate(rows) if keep(r)), key=rank)]


def _write_xlsx(path: str, rows: list) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = (CCAA or "all").title()[:31]
    headers = [h for h, _ in _OUTPUT]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = head_fill, head_font
        cell.alignment = Alignment(vertical="center")

    hi = PatternFill("solid", fgColor="E2EFDA")   # high confidence
    med = PatternFill("solid", fgColor="FFF2CC")  # medium
    conf_idx = headers.index("Email confidence")

    for r in rows:
        ws.append([fn(r) for _h, fn in _OUTPUT])
        c = _confidence(r).lower()
        fill = hi if "high" in c else (med if "medium" in c else None)
        if fill:
            ws.cell(row=ws.max_row, column=conf_idx + 1).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for idx, (h, fn) in enumerate(_OUTPUT, 1):
        longest = max([len(h)] + [len(str(fn(r) or "")) for r in rows] or [0])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(longest + 2, 10), 48)

    wb.save(path)
    return True


def _write_csv(path: str, rows: list) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:  # BOM => Excel UTF-8 clean
        w = csv.writer(f)
        w.writerow([h for h, _ in _OUTPUT])
        for r in rows:
            w.writerow([fn(r) for _h, fn in _OUTPUT])


def main() -> int:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = _fetch(conn)
    if not rows:
        print(f"No qualifying leads at {DB_PATH} (scope={CCAA or 'all'}, tiers={'+'.join(TIERS)}). "
              "If this is unexpected, check you are running on the machine holding the DB.",
              file=sys.stderr)

    out = OUT
    if _write_xlsx(out, rows):
        fmt = "xlsx"
    else:
        out = os.path.splitext(OUT)[0] + ".csv"
        _write_csv(out, rows)
        fmt = "csv (install openpyxl for a formatted .xlsx)"

    n_heir = sum(1 for r in rows if _heir_email(r))
    n_high = sum(1 for r in rows if "high" in _confidence(r).lower())
    n_office = sum(1 for r in rows if _office_email(r))
    n_addr = sum(1 for r in rows if _property_address(r))
    print(f"Exported {len(rows)} leads (tiers {'+'.join(TIERS) or 'all'}, scope={CCAA or 'all'}): "
          f"{n_heir} heir emails ({n_high} high, {n_heir - n_high} medium/other), "
          f"{n_office} office emails, {n_addr} street-level addresses -> {out}  [{fmt}]")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
