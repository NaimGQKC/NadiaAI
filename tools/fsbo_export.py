"""Export FSBO ("venta por particular") listings to an agent-facing Excel.

For a listing agent, an owner already selling without an agent is the warmest
possible prospect for a mandate. This standalone tool scrapes pisos.com's
private-seller filter (see nadia_ai.scrapers.fsbo) and writes a single styled
sheet the agent can open and work top-down — sorted cheapest-first (most
motivated / fastest-moving inventory).

Self-contained: it does not touch nadia_ai.db or the edict pipeline.

Usage: python -m tools.fsbo_export   (writes exports/fsbo_particulares_<date>.xlsx)
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from nadia_ai.scrapers.fsbo import scrape_fsbo

# (header, width, accessor, kind)  — kind drives styling: price|link|wrap|text
COLUMNS = [
    ("Título", 32, lambda l: l.get("title") or "—", "wrap"),
    ("Precio", 13, lambda l: l.get("price_eur") if isinstance(l.get("price_eur"), int) else "—", "price"),
    ("€/m²", 9, lambda l: l.get("eur_per_m2") or "", "text"),
    ("m²", 6, lambda l: l.get("m2") or "", "text"),
    ("Hab.", 6, lambda l: l.get("rooms") or "", "text"),
    ("Motivación", 28, lambda l: l.get("motivation") or "", "wrap"),
    ("Localidad", 18, lambda l: l.get("localidad") or "", "text"),
    ("Dirección", 24, lambda l: l.get("address") or "", "wrap"),
    ("Teléfono", 15, lambda l: l.get("phone") or "—", "text"),
    ("Enlace", 14, lambda l: l.get("listing_url") or "", "link"),
    ("Fecha", 12, lambda l: l.get("scraped_at") or "", "text"),
]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
LINK_FONT = Font(color="2E86C1", underline="single", size=10)
BORDER = Border(bottom=Side(style="thin", color="DDDDDD"))
EUR = Alignment(horizontal="right", vertical="top")


def _sort_key(lead: dict):
    """Cheapest-first; listings without a price sink to the bottom."""
    price = lead.get("price_eur")
    return (0, price) if isinstance(price, int) else (1, 0)


def _write_sheet(ws, leads: list[dict]) -> None:
    for col, (h, w, _, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, lead in enumerate(leads, 2):
        for col, (_, _, accessor, kind) in enumerate(COLUMNS, 1):
            v = accessor(lead)
            c = ws.cell(row=i, column=col, value=v)
            c.border = BORDER
            c.font = Font(bold=(kind == "price"), size=10)
            c.alignment = Alignment(wrap_text=(kind == "wrap"), vertical="top")
            if kind == "price" and isinstance(v, int):
                c.number_format = '#,##0 "€"'
                c.alignment = EUR
            if kind == "link" and v:
                c.value, c.font, c.hyperlink = "Ver anuncio", LINK_FONT, v

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(leads) + 1}"


def build(localidades: list[str] | None = None, max_per_loc: int = 60,
          leads: list[dict] | None = None) -> str:
    """Write the styled FSBO xlsx; returns the output path. Pass pre-scraped
    ``leads`` to avoid re-scraping (the pipeline shares one scrape across the FSBO
    export and the outreach pack)."""
    if leads is None:
        leads = scrape_fsbo(localidades=localidades, max_per_loc=max_per_loc)
    leads = sorted(leads, key=_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "FSBO particulares"
    ws.sheet_properties.tabColor = "27AE60"
    _write_sheet(ws, leads)

    Path("exports").mkdir(exist_ok=True)
    out = f"exports/fsbo_particulares_{date.today().isoformat()}.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    print("Wrote", build())
