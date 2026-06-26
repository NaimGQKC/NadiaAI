"""Generate the per-lead outreach pack — the agent's ready-to-send copy.

Pulls the hottest cohort (FSBO owners with a phone + top open-window inheritance
leads), renders channel-appropriate Spanish outreach for each (see
``nadia_ai.outreach``), and writes ``exports/outreach_<date>.xlsx`` with two
sheets: "Llamadas (FSBO)" (call script + WhatsApp) and "Cartas (Herencia)"
(letters / action notes). The LLM polish (if a key is set) runs ~once per type —
cheap and privacy-safe (placeholders only).

Usage:
    python -m tools.generate_outreach            # deterministic templates
    python -m tools.generate_outreach --llm      # LLM-polished copy
"""

from __future__ import annotations

import sqlite3
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nadia_ai.config import DB_PATH
from nadia_ai.contact_resolve import resolve_for_lead
from nadia_ai.outreach import lead_type, render_outreach
from nadia_ai.outreach_log import init_outreach_log_schema, log_outreach
from nadia_ai.suppression import filter_suppressed, init_suppression_schema
from nadia_ai.tenancy import get_tenant, leads_for_tenant

# Respect NADIA_DB_PATH (the self-hosted persistent DB); the hardcoded default read
# an empty workspace file and broke this step with "no such table: leads".
DB = str(DB_PATH)

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
CALL_FILL = PatternFill("solid", fgColor="D5F5E3")   # FSBO — call now (green)
LETTER_FILL = PatternFill("solid", fgColor="FCF3CF")  # letters (amber)

# (header, width, wrap)
CALL_COLS = [
    ("Tipo", 20, False), ("Inmueble / zona", 24, True), ("Precio", 12, False),
    ("Motivación", 24, True), ("Teléfono", 15, False), ("Guion de llamada", 64, True),
    ("Mensaje WhatsApp", 52, True), ("Notas", 38, True), ("Enlace", 14, False),
]
LETTER_COLS = [
    ("Tipo", 20, False), ("Causante", 26, False), ("Localidad", 16, False),
    ("Canal", 22, False), ("Asunto", 34, True), ("Carta / mensaje", 80, True),
    ("Vía de contacto", 40, True), ("Notas", 40, True), ("Enlace", 14, False),
]


def _inheritance_cohort(conn: sqlite3.Connection, limit: int) -> list[dict]:
    """Top named inheritance/obituary leads, soonest legal deadline first."""
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT * FROM leads
        WHERE (TRIM(COALESCE(causante,'')) <> '' OR TRIM(COALESCE(heir_name,'')) <> '')
        ORDER BY
            CASE WHEN edict_window_days IS NOT NULL AND edict_window_days >= 0 THEN 0 ELSE 1 END,
            edict_window_days ASC,
            first_seen_at DESC
        LIMIT ?
        """,
        (int(limit),),
    ).fetchall()
    return [dict(r) for r in rows]


def _fsbo_cohort(limit: int) -> list[dict]:
    """Fresh FSBO listings with an owner phone (best-effort; network)."""
    try:
        from nadia_ai.scrapers.fsbo import scrape_fsbo

        leads = scrape_fsbo(max_per_loc=limit, enrich_phones=True)
        # Phone-first: a callable owner is the whole point of this sheet.
        leads.sort(key=lambda x: (x.get("phone") is None, x.get("price_eur") or 1e12))
        return leads[:limit]
    except Exception as e:  # noqa: BLE001 — FSBO is a bonus feed, never fail the pack
        import logging

        logging.getLogger("nadia_ai.outreach").warning("FSBO cohort skipped: %s", e)
        return []


def _style_header(ws, cols) -> None:
    for c, (h, w, _) in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{1}"


def _write_row(ws, r: int, values: list, cols, fill) -> None:
    for c, (val, (_, _, wrap)) in enumerate(zip(values, cols), 1):
        cell = ws.cell(r, c, val)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=wrap, vertical="top")
        cell.font = Font(size=10)


def build(limit_inh: int = 30, limit_fsbo: int = 25, use_llm: bool = False,
          db_path: str = DB, fsbo_leads: list[dict] | None = None,
          tenant_id: int | None = None) -> str:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    init_suppression_schema(conn)
    init_outreach_log_schema(conn)
    # Tenant scoping (multi-agent): when tenant_id is given, the inheritance worklist
    # is limited to that agent's territory. Default (None) = full pool, preserving
    # single-agent behaviour. FSBO is configured per-deployment via FSBO_LOCALIDADES,
    # so it is not tenant-scoped here.
    if tenant_id is not None:
        tenant = get_tenant(conn, tenant_id)
        inh = leads_for_tenant(
            conn, tenant,
            extra_where="AND (TRIM(COALESCE(causante,'')) <> '' OR TRIM(COALESCE(heir_name,'')) <> '')",
        )[:limit_inh]
    else:
        inh = _inheritance_cohort(conn, limit_inh)
    # Reuse a pre-scraped FSBO list (the pipeline shares one scrape) or fetch our own.
    fsbo = fsbo_leads if fsbo_leads is not None else _fsbo_cohort(limit_fsbo)
    # Suppression gate: honour every opt-out / do-not-contact BEFORE rendering, so a
    # suppressed person can never reach a worklist regardless of which source resurfaced them.
    inh, _sup_inh = filter_suppressed(conn, inh)
    fsbo, _sup_fsbo = filter_suppressed(conn, fsbo)
    if _sup_inh or _sup_fsbo:
        import logging
        logging.getLogger("nadia_ai.outreach").info(
            "Outreach suppression: removed %d inheritance + %d FSBO leads", _sup_inh, _sup_fsbo
        )

    wb = Workbook()

    # Sheet 1 — FSBO call list.
    ws_call = wb.active
    ws_call.title = "Llamadas (FSBO)"
    _style_header(ws_call, CALL_COLS)
    ws_call.sheet_properties.tabColor = "27AE60"
    r = 2
    for lead in fsbo:
        o = render_outreach(lead, use_llm=use_llm)
        url = lead.get("listing_url") or ""
        _write_row(ws_call, r, [
            o["tipo"], lead.get("address") or lead.get("localidad") or "",
            (f"{lead.get('price_eur'):,}".replace(",", ".") + " €") if lead.get("price_eur") else "",
            lead.get("motivation") or "", lead.get("phone") or "—",
            o["guion_llamada"], o["mensaje"], o["notas"], url,
        ], CALL_COLS, CALL_FILL)
        if url:
            cell = ws_call.cell(r, len(CALL_COLS))
            cell.value, cell.hyperlink = "Ver anuncio", url
            cell.font = Font(color="2E86C1", underline="single", size=10)
        r += 1
    ws_call.auto_filter.ref = f"A1:{get_column_letter(len(CALL_COLS))}{max(r - 1, 1)}"

    # Sheet 2 — inheritance / obituary letters.
    ws_let = wb.create_sheet("Cartas (Herencia)")
    _style_header(ws_let, LETTER_COLS)
    r = 2
    for lead in inh:
        o = render_outreach(lead, use_llm=use_llm)
        import json as _json

        urls = []
        try:
            urls = _json.loads(lead.get("source_urls") or "[]")
        except (_json.JSONDecodeError, TypeError):
            urls = []
        url = urls[0] if urls else ""
        # Resolve HOW to reach this lead — notaría intermediary first, postal floor
        # always. This is the honest contact path, shown in place of a raw office name.
        contact = resolve_for_lead(lead)
        via = f"[{contact.channel}] {contact.target}" if contact else (lead.get("juzgado") or "—")
        _write_row(ws_let, r, [
            o["tipo"], lead.get("causante") or "—", lead.get("localidad") or lead.get("region") or "",
            o["canal"], o["asunto"], o["mensaje"], via, o["notas"], url,
        ], LETTER_COLS, LETTER_FILL)
        if url:
            cell = ws_let.cell(r, len(LETTER_COLS))
            cell.value, cell.hyperlink = "Ver edicto", url
            cell.font = Font(color="2E86C1", underline="single", size=10)
        # Conversion tracking: log the intended outreach (idempotent within cooldown,
        # so re-running the pack the same day doesn't double-count or double-contact).
        if lead.get("id"):
            log_outreach(conn, lead["id"], channel="letter", tipo=o["tipo"])
        r += 1
    ws_let.auto_filter.ref = f"A1:{get_column_letter(len(LETTER_COLS))}{max(r - 1, 1)}"

    conn.close()
    Path("exports").mkdir(exist_ok=True)
    out = f"exports/outreach_{date.today().isoformat()}.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    path = build(use_llm="--llm" in sys.argv)
    print("Wrote", path)
