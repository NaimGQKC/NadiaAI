"""Export the actionable call-list — the agent-facing deliverable.

The DB is a research asset; this turns it into a worklist a listing agent opens and
works top-down. Three honest action classes (not false "all leads are equal"):

  1 · LLAMAR  — notarial *declaración de herederos* (BOE-N) and municipal edicts:
                heirs are actively establishing the succession and the **notaría**
                is named, so there is a warm, legal point of contact today.
  2 · INVESTIGAR — judicial *herencia yacente* (BOE-TEJU) and State succession
                (BOA / BOE-V.B): heirs are unknown/absent but the **court +
                procedure number** are known — a research/monitor lead.
  3 · SEGUIMIENTO — obituaries: a name+city+date death signal only. Wait, then
                check Catastro for property.

Ranking inside a sheet: open legal-deadline first (fewest days left = hottest),
then action class, tier, and recency. Two sheets separate the home market from the
resale inventory: "Zaragoza-Aragón" (what this client works) and "España"
(everything, for referral/resale), plus a "Guía" legend.

Usage: python -m tools.export_call_list   (writes exports/leads_accionables_<date>.xlsx)
"""

from __future__ import annotations

import json
import sqlite3
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB = "nadia_ai.db"
ARAGON = {"Zaragoza", "Huesca", "Teruel"}
NOTARIAL = {"BOE-N", "Tablón", "Tabl�n", "BOA-JD"}          # class 1 — call the notaría/juzgado
JUDICIAL = {"BOE-TEJU", "BOE-V.B"}                               # class 2 — court known, heirs unknown


def _province(lead: dict) -> str:
    """The client-facing province (region rollup already maps towns→province)."""
    return (lead.get("region") or lead.get("localidad") or "").strip()


def _action_class(lead: dict) -> int:
    sub = lead.get("subsource") or ""
    if sub in NOTARIAL:
        return 1
    if sub in JUDICIAL:
        return 2
    return 3


_ACCION = {
    "BOE-N": "Llamar a la notaría; ofrecer servicios a los herederos",
    "Tablón": "Declaración municipal de herederos; contactar juzgado/herederos",
    "BOA-JD": "Sucesión legal Aragón; vigilar inmueble (puede ir a subasta)",
    "BOE-TEJU": "Herencia yacente: investigar inmueble en Catastro y vigilar",
    "BOE-V.B": "Sucesión a favor del Estado; posible subasta — vigilar",
}
_ACCION_DEFAULT = "Fallecimiento reciente: esperar e investigar propiedad (Catastro)"


def _accion(lead: dict) -> str:
    return _ACCION.get((lead.get("subsource") or "").replace("�", "ó"), _ACCION_DEFAULT)


def _heirs(lead: dict) -> str:
    try:
        names = json.loads(lead.get("heir_names_json") or "[]")
    except json.JSONDecodeError:
        names = []
    names = [n for n in names if n]
    return ", ".join(names) if names else "—"


def _contacto(lead: dict) -> str:
    """Compose the reachable office + its phone/email into one cell."""
    office = (lead.get("juzgado") or "").strip()
    phone = (lead.get("contact_phone") or "").strip()
    email = (lead.get("contact_email") or "").strip()
    bits = [b for b in (office, f"Tel. {phone}" if phone else "", email) if b]
    return " · ".join(bits) if bits else "—"


def _ventana(lead: dict) -> str:
    w = lead.get("edict_window_days")
    if w is None:
        return ""
    return f"{w} días restantes" if w > 0 else "cerrada"


def _sort_key(lead: dict):
    w = lead.get("edict_window_days")
    open_window = 0 if (w is not None and w > 0) else 1   # open windows first
    days_left = w if (w is not None and w > 0) else 10**6  # soonest deadline first
    tier_order = {"A": 0, "B": 1, "C": 2, "X": 3}.get(lead.get("tier") or "C", 2)
    dsd = lead.get("days_since_death")
    recency = -(dsd if isinstance(dsd, int) else -1)       # more days since death first
    return (_action_class(lead), open_window, days_left, tier_order, recency)


HEADERS = [
    ("Prioridad", 10), ("Acción recomendada", 46), ("Causante", 30), ("Herederos", 34),
    ("Contacto (notaría/juzgado)", 50), ("Localidad", 16), ("Provincia", 14),
    ("Fecha fallec.", 13), ("Ventana legal", 16), ("Procedimiento", 14),
    ("Tipo", 12), ("Fuente", 14), ("Enlace edicto", 16),
]
TIPO = {1: "Notarial", 2: "Judicial", 3: "Obituario"}
PRIO_LABEL = {1: "1 · LLAMAR", 2: "2 · INVESTIGAR", 3: "3 · SEGUIM."}

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
C1_FILL = PatternFill("solid", fgColor="D5F5E3")   # call now — green
C2_FILL = PatternFill("solid", fgColor="FCF3CF")   # investigate — amber
C3_FILL = PatternFill("solid", fgColor="F2F4F4")   # monitor — grey
LINK_FONT = Font(color="2E86C1", underline="single", size=10)
BORDER = Border(bottom=Side(style="thin", color="DDDDDD"))


def _write_sheet(ws, leads: list[dict]) -> None:
    for col, (h, w) in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    for i, lead in enumerate(leads, 2):
        cls = _action_class(lead)
        fill = {1: C1_FILL, 2: C2_FILL, 3: C3_FILL}[cls]
        urls = json.loads(lead.get("source_urls") or "[]")
        row = [
            PRIO_LABEL[cls], _accion(lead), lead.get("causante") or "—", _heirs(lead),
            _contacto(lead), lead.get("localidad") or "", _province(lead),
            (lead.get("date_of_death") or "")[:10], _ventana(lead),
            lead.get("procedimiento") or "", TIPO[cls],
            (lead.get("subsource") or "").replace("�", "ó"), urls[0] if urls else "",
        ]
        for col, v in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.fill, c.border = fill, BORDER
            c.font = Font(bold=(col == 3), size=10)
            c.alignment = Alignment(wrap_text=(col in (2, 4, 5)), vertical="top")
            if col == len(HEADERS) and v:
                c.value, c.font, c.hyperlink = "Ver edicto", LINK_FONT, v
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(leads) + 1}"


def _write_guide(ws, counts: dict) -> None:
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 92
    title = Font(bold=True, size=14, color="1F3864")
    sect = Font(bold=True, size=12, color="1F3864")
    b = Font(bold=True, size=11)
    n = Font(size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    r = 2
    ws.cell(r, 2, "NadiaAI — Lista de leads accionables").font = title
    r += 1
    ws.cell(r, 2, f"Generado: {date.today().isoformat()}  ·  Aragón {counts['aragon']}  ·  España {counts['total']}").font = n
    r += 2
    ws.cell(r, 2, "PRIORIDADES").font = sect
    r += 1
    for lbl, desc, fill in [
        ("1 · LLAMAR", "Declaración de herederos notarial/municipal: los herederos están "
         "tramitando la sucesión y consta la NOTARÍA. Hay contacto cálido hoy — llama y "
         "ofrece tus servicios.", C1_FILL),
        ("2 · INVESTIGAR", "Herencia yacente judicial o sucesión a favor del Estado: herederos "
         "desconocidos/ausentes, pero consta el JUZGADO y el nº de procedimiento. Investiga el "
         "inmueble en Catastro y vigila.", C2_FILL),
        ("3 · SEGUIMIENTO", "Esquela/obituario: solo señal de fallecimiento (nombre+ciudad+fecha). "
         "Espera por respeto, luego investiga si había propiedad.", C3_FILL),
    ]:
        ws.cell(r, 2, lbl).font = b
        ws.cell(r, 2).fill = fill
        cell = ws.cell(r, 3, desc)
        cell.font, cell.alignment = n, wrap
        ws.row_dimensions[r].height = 48
        r += 1
    r += 1
    ws.cell(r, 2, "ORDEN").font = sect
    r += 1
    ws.cell(r, 3, "Dentro de cada hoja: primero los leads con VENTANA LEGAL abierta (menos días "
                  "restantes = más urgente), luego por prioridad, tier y fecha. La 'ventana legal' "
                  "es el plazo (~30 días) desde la publicación del edicto para que los herederos "
                  "comparezcan: máxima motivación para vender.").font = n
    ws.cell(r, 3).alignment = wrap
    ws.row_dimensions[r].height = 60
    r += 2
    ws.cell(r, 2, "COLUMNAS").font = sect
    r += 1
    for lbl, desc in [
        ("Contacto", "Notaría o juzgado que tramita el caso, con teléfono/email cuando consta en "
                     "el edicto. Es tu punto de entrada legal — NO es el teléfono personal del heredero."),
        ("Herederos", "Personas nombradas en el edicto (cuando las hay). En herencia yacente suelen "
                      "ser desconocidos."),
        ("Ventana legal", "Días restantes del plazo del edicto. 'cerrada' = ya venció (sigue siendo "
                          "lead, menos urgente)."),
        ("Procedimiento", "Nº de expediente notarial/judicial — para referenciar el caso."),
    ]:
        ws.cell(r, 2, lbl).font = b
        cell = ws.cell(r, 3, desc)
        cell.font, cell.alignment = n, wrap
        ws.row_dimensions[r].height = 32
        r += 1


def build(db_path: str = DB) -> str:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    leads = [dict(r) for r in conn.execute("SELECT * FROM leads").fetchall()]
    conn.close()

    # Drop rows with no PERSON to act on. A handling office alone (notaría/juzgado
    # with neither a named deceased nor a named heir) is not a workable lead — the
    # office won't disclose the case, so the agent has no opener. (~40% of the old
    # "actionable" set were these name-less herencia-yacente shells; they polluted
    # Priority 1/2 and produced visible duplicates that dedup couldn't key on.)
    leads = [
        l for l in leads
        if (l.get("causante") or "").strip() or (l.get("heir_name") or "").strip()
    ]
    # Focus the worklist on actionable tiers only. Tier C (insufficient data to
    # act) and Tier X (distress/auction — outreach disabled, buy-side not list-side)
    # are dropped so the agent works a clean A/B list.
    leads = [l for l in leads if (l.get("tier") or "C") in ("A", "B")]
    # Defensive dedup on person+source. merge.py already dedups named leads at write
    # time; this is belt-and-suspenders so the worklist never shows the same case twice.
    seen: set = set()
    deduped = []
    for l in leads:
        key = ((l.get("causante_norm") or l.get("causante") or "").lower().strip(),
               l.get("subsource") or "")
        if key[0] and key in seen:
            continue
        seen.add(key)
        deduped.append(l)
    leads = deduped
    leads.sort(key=_sort_key)
    aragon = [l for l in leads if _province(l) in ARAGON]

    wb = Workbook()
    ws_ar = wb.active
    ws_ar.title = "Zaragoza-Aragón"
    _write_sheet(ws_ar, aragon)
    ws_ar.sheet_properties.tabColor = "27AE60"
    _write_sheet(wb.create_sheet("España (todas)"), leads)
    _write_guide(wb.create_sheet("Guía"), {"aragon": len(aragon), "total": len(leads)})

    Path("exports").mkdir(exist_ok=True)
    out = f"exports/leads_accionables_{date.today().isoformat()}.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    print("Wrote", path)
