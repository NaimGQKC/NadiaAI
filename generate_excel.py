"""Generate the Excel report for Nadia with Tipo column + Leyenda tab."""

import json
import re
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

conn = sqlite3.connect("nadia_ai.db")
conn.row_factory = sqlite3.Row

# Pull leads with edict type + full text for notary extraction
rows = conn.execute("""
    SELECT l.*, e.edict_type, e.source_url as edict_url
    FROM leads l
    LEFT JOIN lead_edicts le ON l.id = le.lead_id
    LEFT JOIN edicts e ON e.id = le.edict_id
    ORDER BY l.tier ASC, l.causante DESC
""").fetchall()

# Deduplicate by lead id
seen = set()
leads = []
for r in rows:
    if r["id"] in seen:
        continue
    seen.add(r["id"])
    leads.append(dict(r))

# Classify tipo
TIPO_MAP = {
    "declaracion_herederos_abintestato": "Declaración de herederos",
    "edicto_judicial_herederos": "Declaración de herederos",
    "sucesion_legal_boa": "Sucesión legal",
}
for l in leads:
    l["tipo"] = TIPO_MAP.get(l.get("edict_type", ""), "Sucesión legal")

# Extract notary from edict text (stored in edicts table via source)
# We'll try to get it from the persons/edicts data
# For now, extract from the edict body if we can get it
notary_cache = {}
edict_rows = conn.execute("""
    SELECT e.id, e.source_url FROM edicts e
""").fetchall()

# Filter noise
noise_words = [
    "causante", "hermanos", "procurador", "ministerio", "diputa",
    "AUTISMO", "diez causantes", "D.ª Agustina",
]
leads = [
    l for l in leads
    if not (l["causante"] and any(n in l["causante"] for n in noise_words))
    and (l["causante"] or l["direccion"])
]

# Clean locality
for l in leads:
    loc = l.get("localidad", "")
    if loc and loc.startswith(("la ", "el ")):
        l["localidad"] = "Zaragoza"

# Sort: A first, then named B, then address-only B
def sort_key(l):
    tier_order = 0 if l["tier"] == "A" else 1
    has_name = 0 if l["causante"] else 1
    return (tier_order, has_name, l["causante"] or "")

leads.sort(key=sort_key)

# === BUILD EXCEL ===
wb = Workbook()

# --- LEADS TAB ---
ws = wb.active
ws.title = "Leads 30-04-2026"

header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
a_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
b_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
link_font = Font(color="2E86C1", underline="single", size=10)
bold_font = Font(bold=True, size=10)
normal_font = Font(size=10)
thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

headers = [
    "Tier", "Tipo", "Causante", "Localidad", "Dirección",
    "Juzgado/Notaría", "Fuente", "Fecha fallec.", "Estado", "Link edicto",
]
widths = [6, 24, 36, 16, 38, 36, 10, 14, 10, 55]
for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(col)].width = w

src_map = {"boa": "BOA", "tablon": "Tablón", "boe_teju": "BOE"}
for i, l in enumerate(leads, 2):
    sources = json.loads(l.get("sources", "[]"))
    source_urls = json.loads(l.get("source_urls", "[]"))
    fill = a_fill if l["tier"] == "A" else b_fill
    vals = [
        l["tier"],
        l["tipo"],
        l["causante"] or "—",
        l.get("localidad", "") or "Zaragoza",
        l.get("direccion", "") or "",
        l.get("juzgado", "") or "",
        ", ".join(src_map.get(s, s) for s in sources),
        l.get("fecha_fallecimiento", "") or "",
        l.get("estado", "Nuevo"),
        source_urls[0] if source_urls else "",
    ]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=col, value=v)
        cell.fill = fill
        cell.border = thin_border
        cell.font = bold_font if col == 3 else normal_font
        if col == 10 and v:
            cell.font = link_font
            cell.hyperlink = v
            cell.value = "Ver edicto"

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{len(leads)+1}"

# --- LEYENDA TAB ---
lg = wb.create_sheet("Leyenda")
lg.sheet_properties.tabColor = "3498DB"

title_font = Font(bold=True, size=14, color="2C3E50")
section_font = Font(bold=True, size=12, color="2C3E50")
bold11 = Font(bold=True, size=11)
normal11 = Font(size=11)
wrap = Alignment(wrap_text=True, vertical="top")

lg.column_dimensions["A"].width = 3
lg.column_dimensions["B"].width = 28
lg.column_dimensions["C"].width = 70

row = 2
lg.cell(row=row, column=2, value="NadiaAI — Guía de leads").font = title_font
row += 2

# Tiers
lg.cell(row=row, column=2, value="TIERS (prioridad)").font = section_font
row += 1
for label, desc, fill in [
    (
        "Tier A",
        "Nombre del causante + dirección del inmueble. "
        "Accionable directamente: puedes buscar la finca en Catastro "
        "y contactar al notario/herederos.",
        a_fill,
    ),
    (
        "Tier B",
        "Nombre del causante O dirección (pero no ambos). "
        "Requiere algo de investigación adicional para localizar "
        "el inmueble o los herederos.",
        b_fill,
    ),
]:
    c2 = lg.cell(row=row, column=2, value=label)
    c2.font = bold11
    c2.fill = fill
    c3 = lg.cell(row=row, column=3, value=desc)
    c3.font = normal11
    c3.alignment = wrap
    lg.row_dimensions[row].height = 45
    row += 1

row += 1
lg.cell(row=row, column=2, value="TIPOS DE EDICTO").font = section_font
row += 1
for label, desc in [
    (
        "Declaración de herederos",
        "Hay herederos pero no había testamento. Un juzgado o notario "
        "está declarando quién hereda. Estos herederos a menudo quieren "
        "vender el inmueble heredado.\n\n"
        "→ Cómo actuar: Busca la notaría que tramita el caso (a veces "
        "aparece en el edicto) y ofrece tus servicios a los herederos.",
    ),
    (
        "Sucesión legal",
        "No se han encontrado herederos. El Gobierno de Aragón está "
        "reclamando el patrimonio. El inmueble acabará en subasta pública.\n\n"
        "→ Cómo actuar: Vigila las subastas en subastas.boe.es "
        "(provincia Zaragoza). Si tiene dirección, puedes informar "
        "a clientes inversores interesados en comprar en subasta.",
    ),
]:
    lg.cell(row=row, column=2, value=label).font = bold11
    c = lg.cell(row=row, column=3, value=desc)
    c.font = normal11
    c.alignment = wrap
    lg.row_dimensions[row].height = 90
    row += 1

row += 1
lg.cell(row=row, column=2, value="FUENTES").font = section_font
row += 1
for label, desc in [
    (
        "Tablón",
        "Tablón de Edictos del Ayuntamiento de Zaragoza. "
        "Publica declaraciones de herederos de la ciudad.",
    ),
    (
        "BOA",
        "Boletín Oficial de Aragón. Publica sucesiones legales "
        "y distribuciones de herencias de toda la comunidad.",
    ),
    (
        "BOE",
        "Boletín Oficial del Estado. Edictos judiciales de ámbito nacional.",
    ),
]:
    lg.cell(row=row, column=2, value=label).font = bold11
    c = lg.cell(row=row, column=3, value=desc)
    c.font = normal11
    c.alignment = wrap
    lg.row_dimensions[row].height = 35
    row += 1

row += 1
lg.cell(row=row, column=2, value="COLUMNAS").font = section_font
row += 1
for label, desc in [
    ("Causante", "Persona fallecida cuyo patrimonio se está tramitando."),
    ("Localidad", "Ciudad/pueblo del fallecido o del inmueble."),
    ("Dirección", "Dirección del inmueble si se ha podido extraer del edicto."),
    (
        "Juzgado/Notaría",
        "Juzgado o notaría que tramita el caso. "
        "Contacta con ellos para ofrecer tus servicios a los herederos.",
    ),
    ("Fecha fallec.", "Fecha de fallecimiento (cuando consta en el edicto)."),
    (
        "Ver edicto",
        "Enlace al documento oficial original. "
        "Ábrelo para ver los detalles completos, incluyendo "
        "la notaría o juzgado que tramita el caso.",
    ),
]:
    lg.cell(row=row, column=2, value=label).font = bold11
    lg.cell(row=row, column=3, value=desc).font = normal11
    row += 1

out = "leads_2026-04-30.xlsx"
wb.save(out)

tier_a = sum(1 for l in leads if l["tier"] == "A")
named = sum(1 for l in leads if l["causante"])
print(f"{len(leads)} leads | {tier_a} Tier A | {named} named")
print(f"File: {out}")
